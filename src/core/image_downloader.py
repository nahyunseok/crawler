import os
import re
import json
import time
import threading
import requests
import pandas as pd
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from PIL import Image
from io import BytesIO
from urllib.parse import urlparse
from src.utils.logger import get_logger
from src.utils.config_manager import allowed_extensions
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 엑셀 셀 하나에 들어갈 수 있는 최대 글자 수 (초과 시 엑셀이 파일을 손상으로 인식)
EXCEL_MAX_CELL_LENGTH = 32000
# PIL 포맷명 → 실제 확장자
PIL_FORMAT_TO_EXT = {
    'JPEG': 'jpg', 'JPG': 'jpg', 'PNG': 'png', 'GIF': 'gif',
    'WEBP': 'webp', 'BMP': 'bmp', 'TIFF': 'tif', 'ICO': 'ico',
}
# 이어받기 기록을 보관하는 폴더 이름 (사이트별로 파일이 하나씩 생긴다)
HISTORY_DIR_NAME = ".history"


class ImageDownloader:
    def __init__(self, config_manager):
        self.logger = get_logger()
        self.config = config_manager
        # 스레드별 requests 세션 보관소 (Session 객체는 스레드 간 공유하지 않는 편이 안전)
        self._thread_local = threading.local()
        # 건너뛴 이유별 집계 (여러 스레드가 함께 올리므로 잠금이 필요하다)
        self._skip_counts = {}
        self._skip_lock = threading.Lock()
        # 도메인·경로가 포함된 쿠키 항목 목록 (이름/값만 담으면 모든 호스트로 전송된다 — _set_cookies 참조)
        self._cookie_items = []

    # ──────────────────────────────────────────────────────────
    # 이어받기(중복 방지) 기록 — 사이트별로 분리 저장
    # ──────────────────────────────────────────────────────────
    @staticmethod
    def _safe_hostname(url):
        """
        호스트명을 파일명으로 쓸 수 있게 정제 (Path Traversal 방지 포함).

        ⛔ 수정금지(DO NOT MODIFY — INTENDED): 점(.)은 허용하되 '..' 는 반드시 죽인다.
        무엇: 연속된 점을 _ 로 바꾸고 앞뒤의 점을 떼어낸다.
        왜:   허용문자에 점이 포함되어 있어서, 주소가 'https://../../etc/passwd' 같은 꼴이면
              netloc 이 '..' 로 뽑혀 그대로 파일명이 됐다. 지금은 뒤에 '.json' 을 붙이기 때문에
              '...json' 이라는 이상한 파일이 되어 실제 탈출은 없지만,
              나중에 이 이름을 '폴더명'으로 쓰는 순간 상위 폴더 탈출이 된다.
              (정상 도메인 example.com 의 점은 그대로 보존되므로 부작용 없음)
        건드리면: 파일명 규칙이 바뀔 때 Path Traversal 취약점으로 되살아난다.
        """
        hostname = urlparse(url).netloc.replace('www.', '').replace(':', '_')
        hostname = re.sub(r'[^a-zA-Z0-9\-\.]', '_', hostname)
        hostname = re.sub(r'\.{2,}', '_', hostname).strip('.')
        return hostname or "unknown"

    @classmethod
    def get_history_path(cls, base_result_dir, source_url):
        """
        사이트별 이어받기 기록 파일 경로.
        ⛔ 예전에는 results/download_history.json 하나에 전 사이트를 몰아 넣어서,
        A사이트를 받고 나면 B사이트 수집에도 영향을 주고 재수집이 아예 불가능했다.
        반드시 '사이트별'로 분리해서 보관한다.
        """
        history_dir = os.path.join(base_result_dir, HISTORY_DIR_NAME)
        os.makedirs(history_dir, exist_ok=True)
        return os.path.join(history_dir, f"{cls._safe_hostname(source_url)}.json")

    @classmethod
    def clear_history(cls, base_result_dir="results"):
        """
        이어받기 기록 전체 삭제 (UI의 '이어받기 기록 초기화' 버튼용).
        Returns: 삭제된 기록 파일 개수
        """
        removed = 0
        history_dir = os.path.join(base_result_dir, HISTORY_DIR_NAME)
        if os.path.isdir(history_dir):
            for name in os.listdir(history_dir):
                if name.endswith(".json"):
                    try:
                        os.remove(os.path.join(history_dir, name))
                        removed += 1
                    except Exception:
                        pass
        # 구버전(전역) 기록 파일도 함께 정리
        legacy = os.path.join(base_result_dir, "download_history.json")
        if os.path.exists(legacy):
            try:
                os.remove(legacy)
                removed += 1
            except Exception:
                pass
        return removed

    def _remove_empty_result_dir(self, save_dir, img_save_dir):
        """
        한 장도 저장되지 않았으면 방금 만든 결과 폴더를 지운다.

        ⛔ 수정금지(DO NOT MODIFY — INTENDED)
        무엇: '이번 수집에서 만든' 빈 폴더만 지운다. 비어 있지 않으면 절대 건드리지 않는다.
        왜:   결과 폴더는 다운로드를 시작하기 전에 미리 만든다. 그런데 필터 때문에 전부
              걸러지면 빈 폴더가 그대로 남았다. 재수집을 반복하면 빈 폴더가 계속 쌓여
              결과 폴더를 열었을 때 어느 것이 실제 결과인지 알 수 없었다(실측 확인).
        건드리면: 사용자의 수집 결과를 실수로 지울 수 있다. 반드시 '비어 있을 때만' 지운다.
        """
        try:
            for path in (img_save_dir, save_dir):
                if os.path.isdir(path) and not os.listdir(path):
                    os.rmdir(path)      # 비어 있지 않으면 OSError 가 나며 아무것도 지우지 않는다
                    self.logger.info(f"Removed empty result folder: {path}")
        except OSError as e:
            # 폴더가 비어있지 않거나 잠겨 있으면 그냥 둔다 (데이터 보존이 우선)
            self.logger.debug(f"Kept result folder ({e})")

    def _note_skip(self, reason):
        """
        이미지를 건너뛴 이유를 집계한다.

        ⛔ 수정금지(DO NOT MODIFY — INTENDED)
        무엇: 다운로드 워커가 이미지를 버릴 때마다 이유를 세어 둔다.
        왜:   예전에는 '이미지 8개 발견 → 7개 저장' 처럼 숫자가 줄어도 이유를 알려주지 않았다.
              사용자는 나머지 1개가 크기 미달인지, 확장자 때문인지, 네트워크 실패인지 알 수 없어
              필터를 어떻게 고쳐야 할지 판단할 수 없었다(침묵 축소).
        건드리면: 수집 결과가 줄어든 이유를 다시 알 수 없게 된다.
        """
        with self._skip_lock:
            self._skip_counts[reason] = self._skip_counts.get(reason, 0) + 1

    def _skip_summary(self):
        """건너뛴 이유를 사람이 읽을 수 있는 한 줄로 만든다. 없으면 빈 문자열."""
        with self._skip_lock:
            if not self._skip_counts:
                return ""
            항목 = sorted(self._skip_counts.items(), key=lambda kv: -kv[1])
        return ", ".join(f"{이유} {개수}개" for 이유, 개수 in 항목)

    def _load_history(self, history_path):
        if not os.path.exists(history_path):
            return set()
        try:
            with open(history_path, 'r', encoding='utf-8') as f:
                return set(json.load(f))
        except Exception:
            return set()

    def _save_history(self, history_path, history):
        """
        이어받기 기록을 저장한다.

        ⛔ 수정금지(DO NOT MODIFY — INTENDED): 설정 파일과 '같은 방식'으로
           임시 파일에 다 쓴 뒤 os.replace 로 원자적으로 교체한다.
        왜: 원본에 직접 쓰다가 프로그램이 강제 종료되면 JSON 이 반쪽만 남아 기록을 못 읽고,
            그러면 이미 받은 이미지를 전부 다시 받는다(수백 MB 재다운로드).
            설정 파일은 이미 이 방식으로 고쳤는데 기록 파일만 예전 방식이라 일관성이 없었다.
        """
        try:
            tmp_path = f"{history_path}.tmp"
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(sorted(history), f, ensure_ascii=False, indent=2)
                f.flush()
                os.fsync(f.fileno())
            os.replace(tmp_path, history_path)
        except Exception as e:
            self.logger.warning(f"Failed to save download history: {e}")

    # ──────────────────────────────────────────────────────────
    # 네트워크 세션 (수동 로그인 쿠키 재사용)
    # ──────────────────────────────────────────────────────────
    def _set_cookies(self, cookies):
        """
        셀레니움에서 받은 쿠키를 requests 가 쓸 수 있는 형태로 보관한다.

        ⛔ 이게 없으면 '수동 로그인'으로 로그인해도 이미지 다운로드는 비로그인 상태로
           요청되어 회원 전용 이미지가 전부 403으로 실패한다.

        ⛔ 수정금지(DO NOT MODIFY — INTENDED): 반드시 '도메인 정보를 살려서' 보관한다.
        무엇: {이름: 값} 딕셔너리가 아니라 도메인·경로가 포함된 쿠키 항목으로 저장한다.
        왜:   예전에는 이름/값만 뽑아 세션에 그대로 넣었다. 그러면 requests 가 그 쿠키를
              '모든 호스트'에 보낸다. 즉 로그인 세션 쿠키가 이미지가 걸려 있는 외부 CDN,
              광고·추적 도메인까지 함께 전송됐다(세션 탈취 위험이 있는 개인정보 유출).
              도메인을 살려 두면 requests 가 스스로 해당 도메인에만 보낸다.
        건드리면: 로그인 쿠키가 제3자 서버로 새어 나간다.
        """
        self._cookie_items = []
        for c in (cookies or []):
            name, value = c.get('name'), c.get('value')
            if not name or value is None:
                continue
            self._cookie_items.append({
                "name": name,
                "value": value,
                # 셀레니움 쿠키의 도메인은 앞에 점이 붙는 경우가 있다(.example.com) — 그대로 둔다.
                "domain": c.get('domain') or '',
                "path": c.get('path') or '/',
            })
        if self._cookie_items:
            도메인들 = sorted({item["domain"] for item in self._cookie_items if item["domain"]})
            self.logger.info(
                f"Reusing {len(self._cookie_items)} login cookies for image download "
                f"(scoped to: {', '.join(도메인들) or 'request host'})"
            )

    def _get_session(self):
        """스레드별 requests 세션 반환 (쿠키 포함)."""
        session = getattr(self._thread_local, "session", None)
        if session is None:
            session = requests.Session()
            session.headers.update({
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
                'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            })
            # 도메인·경로를 지정해 넣으면 requests 가 해당 도메인에만 쿠키를 보낸다
            for item in self._cookie_items:
                try:
                    session.cookies.set(
                        item["name"], item["value"],
                        domain=item["domain"], path=item["path"]
                    )
                except Exception:
                    # 도메인 형식이 이상한 쿠키 하나 때문에 다운로드 전체가 멈추면 안 된다
                    self.logger.debug(f"Skipped malformed cookie: {item['name']}")
            self._thread_local.session = session
        return session

    # ──────────────────────────────────────────────────────────
    # 메인 처리
    # ──────────────────────────────────────────────────────────
    def process_images(self, images_data, base_result_dir="results", progress_callback=None,
                       stop_event=None, cookies=None, message_callback=None):
        """
        이미지를 내려받고 엑셀 리포트를 만든다.

        progress_callback: 진행률(0.0~1.0) 을 받는다.
        message_callback:  사용자에게 보여줄 안내 문구를 받는다.
                           ⛔ 이 통로가 없으면 '몇 개를 왜 건너뛰었는지'가 로그 파일에만 남아
                              사용자는 결과가 줄어든 이유를 알 수 없다(침묵 축소).
        """
        if not images_data:
            self.logger.warning("No images to process.")
            return None

        self._set_cookies(cookies)

        first_img = images_data[0]
        first_source = first_img['source_page']

        # 사이트별 이어받기 기록 로드
        # ⛔ 수정금지(DO NOT MODIFY — INTENDED): 기록은 '항상' 읽고 '항상' 저장한다.
        #    use_resume 는 '건너뛸지 말지'만 결정한다.
        #    예전에는 use_resume 이 꺼져 있으면 빈 기록으로 시작하고 저장도 건너뛰어서,
        #    ① 이번에 받은 것이 기록에 남지 않아 나중에 이어받기를 켜면 또 받았고
        #    ② 빈 기록을 그대로 저장하는 코드로 바뀌면 기존 기록이 통째로 날아갈 위험이 있었다.
        use_resume = self.config.get("use_resume", True)
        history_path = self.get_history_path(base_result_dir, first_source)
        history = self._load_history(history_path)

        # Filter duplicates (Resume feature)
        images_to_process = []
        skipped_count = 0
        for img in images_data:
            if use_resume and img['src'] in history:
                skipped_count += 1
                continue
            images_to_process.append(img)

        if skipped_count > 0:
            self.logger.info(f"중복 제외됨: {skipped_count}개의 이미지는 이미 받아져 건너뜁니다. (이어받기)")

        if not images_to_process:
            self.logger.info("모든 이미지가 이미 다운로드되어 있습니다. (이어받기 완료)")
            if progress_callback: progress_callback(1.0)
            # ⛔ None 을 반환해 호출부가 '새로 저장된 폴더 없음'을 구분할 수 있게 한다.
            #    (예전에는 results 최상위 경로를 돌려줘서 엉뚱한 폴더가 열렸다)
            return None

        # Prepare directory: results/[PageTitle]_[Hostname]_[Date]/
        page_title = first_img.get('page_title', 'Untitled')

        # Sanitize title for filesystem
        safe_title = "".join([c for c in page_title if c.isalnum() or c in (' ', '-', '_')]).strip()
        safe_title = safe_title[:30] # Limit length
        if not safe_title:
            safe_title = "Untitled"

        # 보안 강화: 호스트네임에서 알파벳, 숫자, 점(.), 하이픈(-)만 허용 (Path Traversal 방지)
        hostname = self._safe_hostname(first_source)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # New Format: [Title]_[Domain]_[Time] e.g. "iPhone15Pro_apple.com_20240101..."
        folder_name = f"[{safe_title}]_{hostname}_{timestamp}"

        save_dir = os.path.join(base_result_dir, folder_name)
        img_save_dir = os.path.join(save_dir, "images")
        os.makedirs(img_save_dir, exist_ok=True)

        self.logger.info(f"Saving results to {save_dir}")
        self.logger.info(f"Starting parallel download for {len(images_to_process)} images...")

        downloaded_images = []
        total_images = len(images_to_process)
        completed = 0

        # PRO Feature: Parallel Downloading (5 workers)
        executor = ThreadPoolExecutor(max_workers=5)
        try:
            future_to_img = {
                executor.submit(self._download_single_image, img, idx, img_save_dir, stop_event): img
                for idx, img in enumerate(images_to_process)
            }

            for future in as_completed(future_to_img):
                completed += 1
                if progress_callback:
                    progress_callback(completed / total_images)

                # Reduce log spam by showing summary instead of every single file
                if completed % 20 == 0 or completed == total_images:
                    self.logger.info(f"다운로드 진행 중: {completed}/{total_images} ({int(completed/total_images*100)}%)")

                try:
                    result = future.result()
                except Exception as e:
                    self.logger.debug(f"Download task failed: {e}")
                    result = None

                if result:
                    downloaded_images.append(result)
                    history.add(result['src'])

                if stop_event and stop_event.is_set():
                    # ⛔ 중지 시 '아직 시작되지 않은' 작업들을 즉시 취소한다.
                    #    (예전에는 with 블록이 남은 작업을 전부 기다려서 중지가 먹히지 않았다)
                    self.logger.info("Image downloading stopped by user. Cancelling pending tasks...")
                    for pending in future_to_img:
                        pending.cancel()
                    break
        finally:
            executor.shutdown(wait=False)

        # Save History — 이어받기를 껐더라도 '무엇을 받았는지'는 남긴다 (위 주석 참조)
        self._save_history(history_path, history)

        # 건너뛴 이유 요약을 사용자에게 알린다 (숫자가 줄어든 이유를 반드시 밝힌다)
        요약 = self._skip_summary()
        if 요약:
            self.logger.info(f"Skipped images — {요약}")
            if message_callback:
                message_callback(
                    f"ℹ️ {len(downloaded_images)}개 저장 / {total_images}개 시도 — "
                    f"건너뛴 이유: {요약}"
                )

        if not downloaded_images:
            self.logger.warning("저장된 이미지가 없습니다. (필터 조건 또는 네트워크 확인 필요)")
            self._remove_empty_result_dir(save_dir, img_save_dir)
            return None

        # Generate Excel Report
        self.create_report(downloaded_images, save_dir)
        return save_dir

    # ──────────────────────────────────────────────────────────
    # 개별 다운로드
    # ──────────────────────────────────────────────────────────
    def _fetch_bytes(self, url, referer, stop_event=None):
        """HTTP로 이미지를 받아온다. 용량 상한과 Content-Type을 확인한다."""
        max_bytes = int(self.config.get("max_image_mb", 20)) * 1024 * 1024
        max_retries = 3

        for attempt in range(max_retries):
            if stop_event and stop_event.is_set():
                return None
            try:
                session = self._get_session()
                response = session.get(url, headers={'Referer': referer}, stream=True, timeout=10)

                if response.status_code != 200:
                    self.logger.debug(f"HTTP {response.status_code} for {url[:80]}")
                    response.close()
                else:
                    content_type = (response.headers.get('Content-Type') or '').lower()
                    # 이미지가 아닌 응답(HTML 에러 페이지 등)은 즉시 버린다
                    if content_type and not (content_type.startswith('image/') or 'octet-stream' in content_type):
                        self.logger.debug(f"Not an image ({content_type}): {url[:80]}")
                        response.close()
                        return None

                    # 헤더에 크기가 있으면 먼저 거른다
                    try:
                        declared = int(response.headers.get('Content-Length', 0))
                        if declared and declared > max_bytes:
                            self.logger.debug(f"Too large ({declared} bytes): {url[:80]}")
                            response.close()
                            return None
                    except (TypeError, ValueError):
                        pass

                    # 청크 단위로 받으면서 상한을 넘으면 중단 (메모리 폭탄 방지)
                    buffer = BytesIO()
                    size = 0
                    for chunk in response.iter_content(chunk_size=65536):
                        if stop_event and stop_event.is_set():
                            response.close()
                            return None
                        if not chunk:
                            continue
                        size += len(chunk)
                        if size > max_bytes:
                            self.logger.debug(f"Aborted oversized download: {url[:80]}")
                            response.close()
                            return None
                        buffer.write(chunk)
                    response.close()
                    return buffer.getvalue()

            except Exception as e:
                # We log the shortened URL to prevent flooding logs with giant URLs
                self.logger.debug(f"Request error for {url[:80]}: {e}")

            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)

        return None

    def _download_single_image(self, img, idx, save_dir, stop_event=None):
        """Downloads a single image, supporting both URLs and Data URIs (Base64)."""
        url = img['src']
        stem = f"{idx+1:03d}_{img['filename']}"

        image_content = None

        # --- Step 1: Retrieve Content ---
        if url.startswith('data:image/'):
            # Support for embedded Base64 images
            try:
                import base64
                if ',' not in url:
                    return None
                header, encoded = url.split(',', 1)
                if ';base64' not in header:
                    return None
                image_content = base64.b64decode(encoded)
            except Exception as e:
                self.logger.debug(f"Failed to decode base64 for {stem}: {e}")
                self._note_skip("페이지 내장 이미지 해독 실패")
                return None
        else:
            image_content = self._fetch_bytes(url, img.get('source_page', url), stop_event)

        if not image_content:
            if not (stop_event and stop_event.is_set()):
                self._note_skip("내려받기 실패(네트워크·차단·형식)")
            return None

        # --- Step 2: Validate, Save & Process ---
        try:
            # Check constraints
            image = Image.open(BytesIO(image_content))
            image.verify()  # 깨진 파일 조기 차단
            image = Image.open(BytesIO(image_content))  # verify() 후에는 재오픈 필요

            min_width = self.config.get("min_width", 0)
            min_height = self.config.get("min_height", 0)

            if image.width < min_width or image.height < min_height:
                self.logger.debug(f"Skipped {stem}: Too small ({image.width}x{image.height})")
                self._note_skip(f"최소 크기({min_width}x{min_height}) 미달")
                return None

            # ⛔ 확장자는 '실제 이미지 포맷' 기준으로 붙인다.
            #    예전에는 URL에 확장자가 없으면 무조건 .jpg 를 붙여 PNG가 .jpg로 저장됐다.
            ext = PIL_FORMAT_TO_EXT.get((image.format or '').upper(), 'jpg')

            # ⛔ 수정금지(DO NOT MODIFY — INTENDED): 실제 포맷도 허용 확장자 목록으로 검사한다.
            # 왜: 크롤러 단계의 검사는 '주소에 확장자가 있을 때'만 동작한다. 그래서
            #     /photo?id=1 처럼 확장자 없는 주소는 필터를 그냥 통과했고, 내려받아 보니
            #     GIF 인데도 GIF 체크를 껐는지 여부와 무관하게 .gif 로 저장됐다.
            #     화면의 체크박스와 실제로 저장되는 파일이 어긋나던 원인이다.
            allowed = allowed_extensions(self.config)
            if f".{ext}" not in allowed:
                self.logger.debug(f"Skipped {stem}: format .{ext} not in allowed extensions {allowed}")
                self._note_skip(f"허용하지 않는 형식(.{ext})")
                return None

            filename = f"{stem}.{ext}"
            filepath = os.path.join(save_dir, filename)

            # Write raw image data
            with open(filepath, 'wb') as f:
                f.write(image_content)

            # Save supporting text metadata
            txt_filepath = os.path.join(save_dir, f"{stem}.txt")
            try:
                with open(txt_filepath, "w", encoding="utf-8") as f:
                    f.write(f"파일이름: {filename}\n")
                    # Prevent bloating txt file if it's a giant base64 string
                    clean_src = url if len(url) < 200 else "Base64 Data (String too long for log)"
                    f.write(f"출처링크: {clean_src}\n")
                    f.write(f"이미지설명: {img.get('description', '없음')}\n")
                    f.write(f"단락 제목(Heading): {img.get('heading', '없음')}\n")
                    f.write("-" * 40 + "\n")
                    f.write(f"[주변 텍스트 문맥]\n")
                    f.write(f"{img.get('context', '없음')}\n")
            except Exception:
                pass # Non-critical failure

            # Add tracking metadata
            img['saved_filename'] = filename
            img['resolution'] = f"{image.width}x{image.height}"
            self.logger.debug(f"Successfully saved: {filename}")
            return img

        except Exception as e:
            self.logger.debug(f"Processing failed for {stem}: {e}")
            self._note_skip("이미지가 아니거나 파일이 손상됨")
            return None

    # ──────────────────────────────────────────────────────────
    # 엑셀 리포트
    # ──────────────────────────────────────────────────────────
    def create_report(self, images_data, output_dir):
        """Creates an Excel report from value data."""
        if not images_data:
            return

        df = pd.DataFrame(images_data)

        # Select and Reorder columns
        columns = ['saved_filename', 'heading', 'description', 'context', 'src', 'resolution', 'source_page']
        for col in columns:
            if col not in df.columns:
                df[col] = ""

        # Add scrape time
        df['scrape_time'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        columns.append('scrape_time')

        df = df[columns].copy()

        # ⛔ 엑셀은 셀 하나에 32,767자를 넘으면 '파일 손상'으로 인식한다.
        #    Base64 이미지 주소는 수만 자가 되므로 반드시 잘라서 넣는다.
        def _shorten(value):
            text = str(value) if value is not None else ""
            if len(text) > EXCEL_MAX_CELL_LENGTH:
                return text[:EXCEL_MAX_CELL_LENGTH] + " ...(생략됨)"
            return text

        for col in df.columns:
            df[col] = df[col].map(_shorten)

        df.columns = [
            '수집된 파일명',
            '소속 문단 제목 (주제 파악용)',
            '이미지 자체 설명 (Alt/Title 등)',
            '주변 텍스트 문맥 (본문 내용)',
            '실제 다운로드 통로 URL',
            '해상도',
            '출처 페이지 사이트',
            '수집 일시'
        ]

        excel_path = os.path.join(output_dir, "checklist.xlsx")
        try:
            # Save using Pandas
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='수집 결과')

                # Format using openpyxl
                worksheet = writer.sheets['수집 결과']

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_alignment = Alignment(horizontal="center", vertical="center")
                wrap_alignment = Alignment(wrap_text=True, vertical="center")

                # Format Headers
                for col_num, cell in enumerate(worksheet[1], 1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    cell.border = thin_border

                # Format Data Cells & Auto-width
                column_widths = {'A': 25, 'B': 35, 'C': 40, 'D': 50, 'E': 40, 'F': 15, 'G': 40, 'H': 20}
                for col_letter, width in column_widths.items():
                    worksheet.column_dimensions[col_letter].width = width

                for row in worksheet.iter_rows(min_row=2):
                    for idx, cell in enumerate(row):
                        cell.border = thin_border
                        if idx in [1, 2, 3]: # Heading, Description and Context -> Wrap text
                            cell.alignment = wrap_alignment
                        else:
                            cell.alignment = Alignment(vertical="center")

            self.logger.info(f"Excel report saved: {excel_path}")
        except Exception as e:
            self.logger.error(f"Failed to save Excel report: {e}")
